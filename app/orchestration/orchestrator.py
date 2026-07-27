"""
Orchestrator — wires the three agents together and schedules them.
Runs the MS Agent framework pipeline on market hours.
"""

# load_dotenv() must run before importing modules that read env at import time
# (e.g. agents.alert.alert_agent), so imports intentionally follow it.
# ruff: noqa: E402

import argparse
import csv
import json
import os
import tempfile
import time
import traceback
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from dotenv import load_dotenv

load_dotenv()

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger

from app.agents.analyst.analyst_agent import AnalystAgent, recommendation
from app.agents.alert.alert_agent import AlertAgent
from app.agents.extraction.extraction_agent import ExtractionAgent
from app.agents.trader.trader_agent import TraderAgent
from app.services.portfolio_service import PortfolioService
from app.services.trader_service import TraderService
from app.workflows.momentum import build_momentum_pipeline
from app.agents.scanner.scanner_agent import (
    ScannerAgent,
    load_watchlist,
    load_source_map,
)
from app.agents.scanner.scan_history import (
    get_fresh_breakouts,
    get_new_tickers,
    load_history,
    save_history,
)
from app.agents.scanner.market_cycle import get_market_cycle_context
from app.agents.scanner.market_narrative import (
    build_deterministic_narrative,
    save_market_narrative,
)
from app.agents.scanner.sector_allocation import (
    compute_portfolio_sector_weights,
    compute_sector_prevalence,
    latest_snapshot,
    load_sector_history,
    save_sector_snapshot,
    with_deltas,
)
from app.schemas import StockRecord
from app.core.config import (
    ANALYSIS_JSON,
    ANALYSIS_XLSX,
    PIPELINE_RUN_TIMEOUT_SECONDS,
    PIPELINE_RUNS_CSV,
    PIPELINE_STALE_GRACE_SECONDS,
    PIPELINE_STATUS_JSON,
    PORTFOLIO_VALUE_CSV,
    SCAN_RESULTS_JSON,
)
from app.repositories.notifications_repo import build_notifications_repository
from app.repositories.pipeline_status_repo import (
    PipelineRunActiveError,
    PipelineRunInactiveError,
    PipelineStatusRepository,
)
from app.schemas.analysis_artifact import build_analysis_payload
from app.schemas.notification import NotificationCategory, NotificationSeverity
from app.schemas.pipeline_status import PipelineStage, PipelineState, StageState
from app.schemas.source_health import (
    SourceHealth,
    SourceName,
    SourceState,
    derive_pipeline_outcome,
)
from app.workflows.pipeline import PipelineStepEvent


SCAN_OUTPUT = SCAN_RESULTS_JSON
RUN_LOG = PIPELINE_RUNS_CSV
PORTFOLIO_VALUE_LOG = PORTFOLIO_VALUE_CSV
_RUN_LOG_FIELDS = [
    "run_id",
    "start",
    "end",
    "duration_seconds",
    "scanned",
    "analysed",
    "buy_alerts",
    "sell_alerts",
    "actionable",
    "sources",
    "source_health_json",
    "status",
    "errors",
]
_SOURCE_COMMENTS: dict[str, str] = {
    "ww_extraction": "WhaleWisdom heat map – institutional filer top holdings",
    "vcp_screener": "Minervini pure VCP setup – S&P 500 screened via FMP API",
    "tv_screener": "TradingView screener – Stage 2 pre-filter (price>SMA200, SMA50>SMA150, within 35% of 52w high)",
    "tv_screener_uk": "TradingView screener – LSE Stage 2 pre-filter (price>SMA200, SMA50>SMA150, within 35% of 52w high)",
}
ANALYSIS_OUTPUT = ANALYSIS_JSON
EXCEL_OUTPUT = ANALYSIS_XLSX
MARKET_OPEN_HOUR = 9
MARKET_OPEN_MIN = 30
MARKET_CLOSE_HOUR = 16
MARKET_CLOSE_MIN = 0


def _append_portfolio_snapshot(total_value: float, total_cost: float) -> None:
    """Append one portfolio value snapshot (timestamped) to PORTFOLIO_VALUE_LOG."""
    log_path = Path(PORTFOLIO_VALUE_LOG)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not log_path.exists() or log_path.stat().st_size == 0
    with open(log_path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh, fieldnames=["timestamp", "total_value", "total_cost"]
        )
        if write_header:
            writer.writeheader()
        writer.writerow(
            {
                "timestamp": datetime.now(timezone.utc).isoformat(timespec="minutes"),
                "total_value": round(total_value, 2),
                "total_cost": round(total_cost, 2),
            }
        )


def _migrate_run_log_header(log_path: Path) -> None:
    """Rewrite an existing run log onto the current field set, in place.

    Older logs may be missing columns (``run_id``, ``source_health_json``)
    added by #40/#42. Rows are preserved; missing fields read back as empty
    strings so legacy history keeps displaying.
    """
    with open(log_path, newline="", encoding="utf-8") as existing:
        reader = csv.DictReader(existing)
        existing_fields = reader.fieldnames or []
        if existing_fields == _RUN_LOG_FIELDS:
            return
        prior_rows = [dict(row) for row in reader]
    fd, temporary_name = tempfile.mkstemp(
        dir=log_path.parent, prefix=f".{log_path.name}.", suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", newline="", encoding="utf-8") as migrated:
            writer = csv.DictWriter(migrated, fieldnames=_RUN_LOG_FIELDS)
            writer.writeheader()
            for row in prior_rows:
                writer.writerow({key: row.get(key, "") for key in _RUN_LOG_FIELDS})
            migrated.flush()
            os.fsync(migrated.fileno())
        os.replace(temporary_name, log_path)
    finally:
        Path(temporary_name).unlink(missing_ok=True)


def _append_run_log(entry: dict) -> None:
    """Append one run record to the CSV log, writing the header if the file is new."""
    log_path = Path(RUN_LOG)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    if log_path.exists() and log_path.stat().st_size:
        _migrate_run_log_header(log_path)
    write_header = not log_path.exists() or log_path.stat().st_size == 0
    with open(log_path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_RUN_LOG_FIELDS)
        if write_header:
            writer.writeheader()
        writer.writerow({k: entry.get(k, "") for k in _RUN_LOG_FIELDS})


#: How each terminal pipeline state renders as a refresh notification (#80).
_REFRESH_NOTIFICATION: dict[PipelineState, tuple[str, NotificationSeverity, str]] = {
    PipelineState.COMPLETE: (
        "refresh_complete",
        NotificationSeverity.INFO,
        "Data refresh complete",
    ),
    PipelineState.PARTIAL: (
        "refresh_partial",
        NotificationSeverity.WARNING,
        "Data refresh partially complete",
    ),
    PipelineState.FAILED: (
        "refresh_failed",
        NotificationSeverity.ERROR,
        "Data refresh failed",
    ),
}


def _emit_run_notifications(
    terminal_state: PipelineState,
    source_health: dict[SourceName, SourceHealth],
    *,
    run_id: str,
    scanned: int,
    analysed: int,
    buy_alerts: int,
    sell_alerts: int,
    duration: float,
) -> None:
    """Record refresh + source-health notifications for one finished run (#80).

    Best-effort: a notification-store failure must never break the pipeline,
    so everything here is wrapped. Only genuine source *failures* are surfaced
    (empty/skipped are routine and already shown in the status bar), keeping
    the feed signal-rich.
    """
    try:
        repo = build_notifications_repository()
        mapping = _REFRESH_NOTIFICATION.get(terminal_state)
        if mapping is not None:
            event_type, severity, title = mapping
            repo.record(
                NotificationCategory.REFRESH,
                event_type,
                title,
                severity=severity,
                body=(
                    f"Scanned {scanned}, analysed {analysed}"
                    f" — {buy_alerts} buy / {sell_alerts} sell"
                    f" alert(s) in {duration:g}s."
                ),
                run_id=run_id,
            )
        for health in source_health.values():
            if health.state is not SourceState.FAILED:
                continue
            repo.record(
                NotificationCategory.SOURCE,
                "source_failed",
                f"Source failed — {health.source.label}",
                severity=NotificationSeverity.ERROR,
                body=health.display_message or "This data source failed.",
                run_id=run_id,
            )
    except Exception as error:  # pragma: no cover - defensive
        print(f"[notify] run notification error: {error}")


def _cached_extraction_health(
    source_map: dict[str, tuple[bool, bool]],
) -> dict[SourceName, SourceHealth]:
    """Return SKIPPED health for extraction sources reused from a cached run.

    When ``pipeline(extract=False)`` runs, WhaleWisdom/StockTwits were not
    refreshed this run; their counts come from the cached extraction file so
    they must not be reported as a freshly-successful "ok" source.
    """
    stocktwits = sum(stocktwits for stocktwits, _ in source_map.values())
    whale_wisdom = sum(whale for _, whale in source_map.values())
    return {
        SourceName.WHALE_WISDOM: SourceHealth(
            source=SourceName.WHALE_WISDOM,
            state=SourceState.SKIPPED,
            count=whale_wisdom,
            detail_code="cached_input",
            display_message="Using cached WhaleWisdom input; this source was not refreshed.",
        ),
        SourceName.STOCKTWITS: SourceHealth(
            source=SourceName.STOCKTWITS,
            state=SourceState.SKIPPED,
            count=stocktwits,
            detail_code="cached_input",
            display_message="Using cached StockTwits input; this source was not refreshed.",
        ),
    }


_SEPA_LABELS = {
    "above_150_200": "P>SMA150&200",
    "sma150_above_200": "SMA150>200",
    "sma200_rising": "SMA200 Rising",
    "sma50_above_150_200": "SMA50>150&200",
    "above_25pct_of_low": ">=25% abv Low",
    "within_25pct_of_high": "<=25% frm High",
    "rs_leader": "RS Leader",
    "above_sma50": "P>SMA50",
}
_SEPA_KEYS = list(_SEPA_LABELS.keys())

_HEADERS = [
    "Ticker",
    "Market",
    "Source",
    "StockTwits",
    "Whale Wisdom",
    "Score",
    "CANSLIM",
    "Momentum",
    "Stage",
    "Zone",
    "Prev Entry",
    "Next Entry",
    "Stop",
    "Risk %",
    "1R Target",
    "2R Target",
    "3R Target",
    "R:R",
    "Price",
    "P/E",
    "RSI",
    "Rel Vol",
    "% 52w High",
    "% Chg Week",
    "Rel Str vs SPY",
    "EPS Growth",
    "ROE",
    "Inst Ownership %",
    "Inst Count",
    "WW Buyers",
    "WW Sellers",
    "WW Net",
    "Congress Buys",
    "Congress Sells",
    "Congress Net",
    "Senate Buys",
    "Senate Sells",
    "Senate Net",
    "SPY Uptrend",
    "SMA Stack",
    "SMA50",
    "Near High",
    "RSI Zone",
    "Vol Zone",
    "SEPA",
    *[f"SEPA: {v}" for v in _SEPA_LABELS.values()],
    "Sector",
    "VCP Breakout",
    "MYB Breakout",
    "MYB Pivot",
    "Recommendation",
    "As Of",
    "Summary",
]

# Derive column positions dynamically so they stay correct as columns are added.
_SIGNAL_COL_START = _HEADERS.index("SMA Stack") + 1  # 1-indexed
_SIGNAL_COL_END = _HEADERS.index("Vol Zone") + 1
_SEPA_COL_START = _HEADERS.index("SEPA") + 1
_SEPA_COL_END = _SEPA_COL_START + len(
    _SEPA_KEYS
)  # summary + 8 conditions (exclusive end = summary+8)

_SCORE_FILLS = {
    "high": PatternFill("solid", fgColor="C6EFCE"),  # green  ≥8
    "mid": PatternFill("solid", fgColor="FFEB9C"),  # yellow 6–7
    "low": PatternFill("solid", fgColor="FFC7CE"),  # red    ≤5
}
_SIGNAL_FILLS = {
    "+": PatternFill("solid", fgColor="C6EFCE"),  # green  — positive for buying
    "-": PatternFill("solid", fgColor="FFC7CE"),  # red    — negative / caution
    "~": PatternFill("solid", fgColor="FFEB9C"),  # yellow — neutral
}
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NEAR_ENTRY_FILL = PatternFill("solid", fgColor="E2EFDA")


def _yahoo_url(ticker: str) -> str:
    return f"https://finance.yahoo.com/quote/{ticker.replace('.', '-')}"


def _score_fill(score: int) -> PatternFill:
    if score >= 8:
        return _SCORE_FILLS["high"]
    if score >= 6:
        return _SCORE_FILLS["mid"]
    return _SCORE_FILLS["low"]


def _pct(value: float | None) -> str:
    return f"{value * 100:.1f}%" if value is not None else ""


def _signals(r: StockRecord) -> list[str]:
    """Return 5 buy-signal indicators (+/-/~) from scan data.

    These correspond to the conditions previously used in the heuristic score
    and are shown as coloured columns in the Excel output.
    """
    sma150 = r.sma150
    sma200 = r.sma200
    sma50 = r.sma50
    rsi = r.rsi14
    pct = r.pct_from_52w_high
    vol = r.rel_volume

    # SMA Stack: price > SMA150 > SMA200 → bullish macro structure
    if sma150 and sma200 and r.price > sma150 > sma200:
        sma_stack = "+"
    elif sma150 and sma200 and r.price < sma150 and r.price < sma200:
        sma_stack = "-"
    else:
        sma_stack = "~"

    # SMA50: short-term trend
    sma50_sig = "+" if sma50 and r.price > sma50 else "-"

    # Near High: -15% to -1% is the ideal consolidation zone; < -30% is too extended
    if pct is not None and -15 <= pct <= -1:
        near_high = "+"
    elif pct is not None and pct < -30:
        near_high = "-"
    else:
        near_high = "~"

    # RSI Zone: 50-80 healthy, >80 overbought, <40 weak, 40-50 neutral
    if rsi is not None and 50 <= rsi <= 80:
        rsi_zone = "+"
    elif rsi is not None and (rsi > 80 or rsi < 40):
        rsi_zone = "-"
    else:
        rsi_zone = "~"

    # Vol Zone: elevated volume signals institutional participation
    if vol >= 1.5:
        vol_zone = "+"
    elif vol < 0.7:
        vol_zone = "-"
    else:
        vol_zone = "~"

    return [sma_stack, sma50_sig, near_high, rsi_zone, vol_zone]


def _record_to_row(
    r: StockRecord, portfolio_tickers: set[str] | None = None
) -> list[object]:
    a = r.analysis
    if not a:
        return [r.ticker] + [""] * (len(_HEADERS) - 1)
    risk = (
        f"{(a.entry_price - a.stop_loss) / a.entry_price * 100:.1f}%"
        if a.entry_price and a.stop_loss
        else ""
    )
    tmpl = a.sepa_template or {}
    sepa_vals = [tmpl.get(k, False) for k in _SEPA_KEYS]
    sepa_summary = f"{sum(sepa_vals)}/8"

    rec = recommendation(a)
    if rec == "Avoid" and portfolio_tickers and r.ticker in portfolio_tickers:
        rec = "SELL"

    return [
        r.ticker,
        "UK" if r.ticker.endswith(".L") else "US",
        r.source,
        "Y" if r.in_stocktwits else "",
        "Y" if r.in_whale_wisdom else "",
        f"{a.score}/10",
        f"{a.canslim.total}/14" if a.canslim else "",
        f"{a.momentum.total}/14" if a.momentum else "",
        a.stage,
        a.entry_zone,
        a.prev_entry_price,
        a.entry_price,
        a.stop_loss,
        risk,
        a.r_multiples.get("1.0R") if a.r_multiples else "",
        a.r_multiples.get("2.0R") if a.r_multiples else "",
        a.r_multiples.get("3.0R") if a.r_multiples else "",
        f"{a.reward_risk_ratio:.1f}x" if a.reward_risk_ratio else "",
        r.price,
        round(r.pe_ratio, 1) if r.pe_ratio else "",
        r.rsi14,
        r.rel_volume,
        r.pct_from_52w_high,
        r.pct_change_week,
        r.rel_strength_vs_spy,
        _pct(r.eps_growth),
        _pct(r.roe),
        _pct(r.inst_ownership_pct),
        r.inst_count if r.inst_count is not None else "",
        r.funds_buying if r.funds_buying is not None else "",
        r.funds_selling if r.funds_selling is not None else "",
        r.funds_net if r.funds_net is not None else "",
        r.congress_buys if r.congress_buys is not None else "",
        r.congress_sells if r.congress_sells is not None else "",
        (r.congress_buys or 0) - (r.congress_sells or 0)
        if r.congress_buys is not None or r.congress_sells is not None
        else "",
        r.senate_buys if r.senate_buys is not None else "",
        r.senate_sells if r.senate_sells is not None else "",
        (r.senate_buys or 0) - (r.senate_sells or 0)
        if r.senate_buys is not None or r.senate_sells is not None
        else "",
        "Y" if r.spy_uptrend else "N",
        *_signals(r),
        sepa_summary,
        *["Y" if v else "N" for v in sepa_vals],
        r.sector or "",
        "Y" if a.fresh_breakout else "",
        "Y" if a.multiyear_breakout else "",
        a.multiyear_pivot if a.multiyear_pivot else "",
        rec,
        r.as_of,
        a.summary,
    ]


def _apply_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.append(_HEADERS)
    for col, _ in enumerate(_HEADERS, 1):
        cell = ws.cell(1, col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _set_col_widths(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    widths = {
        "A": 8,  # Ticker
        "B": 18,  # Source
        "C": 11,  # StockTwits
        "D": 12,  # Whale Wisdom
        "E": 8,  # Score
        "F": 10,  # CANSLIM
        "G": 10,  # Momentum
        "H": 10,  # Stage
        "I": 8,  # Zone
        "J": 11,  # Prev Entry
        "K": 11,  # Next Entry
        "L": 10,  # Stop
        "M": 8,  # Risk %
        "N": 10,  # 1R Target
        "O": 10,  # 2R Target
        "P": 10,  # 3R Target
        "Q": 7,  # R:R
        "R": 10,  # Price
        "S": 7,  # P/E
        "T": 7,  # RSI
        "U": 8,  # Rel Vol
        "V": 11,  # % 52w High / % Chg Week
        "W": 14,  # Rel Str vs SPY
        "X": 11,  # EPS Growth
        "Y": 8,  # ROE
        "Z": 15,  # Inst Ownership
        "AA": 10,  # Inst Count
        "AB": 9,  # WW Buyers
        "AC": 9,  # WW Sellers
        "AD": 8,  # WW Net
        "AE": 12,  # Congress Buys
        "AF": 12,  # Congress Sells
        "AG": 10,  # Congress Net
        "AH": 10,  # Senate Buys
        "AI": 10,  # Senate Sells
        "AJ": 10,  # Senate Net
        "AK": 10,  # SPY Uptrend
        "AL": 9,  # SMA Stack
        "AM": 7,  # SMA50
        "AN": 9,  # Near High
        "AO": 8,  # RSI Zone
        "AP": 8,  # Vol Zone
        "AQ": 7,  # SEPA summary
        "AR": 13,
        "AS": 13,
        "AT": 13,
        "AU": 13,  # SEPA conditions 1-4
        "AV": 13,
        "AW": 13,
        "AX": 13,
        "AY": 13,  # SEPA conditions 5-8
        "AZ": 12,  # Sector
        "BA": 12,  # Recommendation
        "BB": 12,  # As Of
        "BC": 50,  # Summary
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _format_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    data_row: int,
    r: StockRecord,
) -> None:
    """Apply fills, hyperlink, and alignment for a single data row."""
    score = r.analysis.score if r.analysis else 0
    fill = _score_fill(score)
    link_cell = ws.cell(data_row, 1)
    assert isinstance(link_cell, Cell)
    link_cell.hyperlink = Hyperlink(ref="", target=_yahoo_url(r.ticker))
    link_cell.font = Font(color="0563C1", underline="single")
    link_cell.fill = fill
    ws.cell(data_row, 4).fill = fill
    if r.analysis and r.analysis.entry_zone in ("broken_out", "approaching"):
        ws.cell(data_row, 8).fill = _NEAR_ENTRY_FILL

    for col in range(_SIGNAL_COL_START, _SIGNAL_COL_END + 1):
        cell = ws.cell(data_row, col)
        if cell.value in _SIGNAL_FILLS:
            cell.fill = _SIGNAL_FILLS[cell.value]
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(_SEPA_COL_START + 1, _SEPA_COL_END + 1):
        cell = ws.cell(data_row, col)
        if cell.value == "Y":
            cell.fill = _SIGNAL_FILLS["+"]
        elif cell.value == "N":
            cell.fill = _SIGNAL_FILLS["-"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sepa_cell = ws.cell(data_row, _SEPA_COL_START)
    try:
        count = int(str(sepa_cell.value).split("/")[0])
    except (ValueError, AttributeError):
        count = 0
    if count == 8:
        sepa_fill = _SIGNAL_FILLS["+"]
    elif count >= 6:
        sepa_fill = _SIGNAL_FILLS["~"]
    else:
        sepa_fill = _SIGNAL_FILLS["-"]
    sepa_cell.fill = sepa_fill
    sepa_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_name in ("VCP Breakout", "MYB Breakout"):
        col = _HEADERS.index(col_name) + 1
        cell = ws.cell(data_row, col)
        if cell.value == "Y":
            cell.fill = _SIGNAL_FILLS["+"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    market_col = _HEADERS.index("Market") + 1
    market_cell = ws.cell(data_row, market_col)
    if market_cell.value == "UK":
        market_cell.fill = PatternFill("solid", fgColor="DBEAFE")  # pale blue
    market_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(_HEADERS) + 1):
        ws.cell(data_row, col).alignment = Alignment(vertical="center")


_NEW_TICKER_FILL = PatternFill("solid", fgColor="FFF2CC")  # pale yellow — new tickers
_BREAKOUT_FILL = PatternFill("solid", fgColor="FFE0B2")  # pale orange — fresh breakouts


def _write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    records: list[StockRecord],
    held: set[str],
    ranked: bool = False,
) -> None:
    """Populate a worksheet with header + formatted data rows."""
    _apply_header(ws)
    _set_col_widths(ws)
    for rank, r in enumerate(records, 1):
        row = _record_to_row(r, held)
        if ranked:
            row[0] = f"{rank}. {r.ticker}"
        ws.append(row)
        _format_row(ws, ws.max_row, r)


def write_excel(
    records: list[StockRecord],
    path: str | Path,
    portfolio_tickers: set[str] | None = None,
    new_tickers: set[str] | None = None,
) -> None:
    """Write analysis results to a formatted Excel workbook."""
    held = portfolio_tickers or set()
    new = new_tickers or set()
    wb = openpyxl.Workbook()

    # --- Sheet 1: All results ---
    ws_all = wb.active
    ws_all.title = "All Results"
    _write_sheet(ws_all, records, held)

    # Highlight new tickers with a pale-yellow row tint on Sheet 1
    for row_idx in range(2, ws_all.max_row + 1):
        ticker_cell = ws_all.cell(row_idx, 1)
        raw = str(ticker_cell.value or "").lstrip("0123456789. ")
        if raw in new:
            for col in range(1, len(_HEADERS) + 1):
                cell = ws_all.cell(row_idx, col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _NEW_TICKER_FILL

    # --- Sheet 2: New Tickers ---
    new_records = [r for r in records if r.ticker in new]
    if new_records:
        ws_new = wb.create_sheet("New Tickers")
        _write_sheet(ws_new, new_records, held)
        # All rows on this sheet get the yellow tint
        for row_idx in range(2, ws_new.max_row + 1):
            for col in range(1, len(_HEADERS) + 1):
                cell = ws_new.cell(row_idx, col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _NEW_TICKER_FILL

    # --- Sheet 3: Fresh Breakouts ---
    breakout_records = [r for r in records if r.analysis and r.analysis.fresh_breakout]
    if breakout_records:
        ws_bo = wb.create_sheet("Breakouts")
        _write_sheet(ws_bo, breakout_records, held)
        for row_idx in range(2, ws_bo.max_row + 1):
            for col in range(1, len(_HEADERS) + 1):
                cell = ws_bo.cell(row_idx, col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _BREAKOUT_FILL

    # --- Sheet 4: Top 20 ---
    top20 = [r for r in records if r.analysis][:20]
    ws_top = wb.create_sheet("Top 20")
    _write_sheet(ws_top, top20, held, ranked=True)

    wb.save(path)
    parts = [f"{len(records)} records"]
    if new_records:
        parts.append(f"{len(new_records)} new")
    if breakout_records:
        parts.append(f"{len(breakout_records)} breakouts")
    print(f"      Excel saved -> {path}  ({', '.join(parts)})")


def is_market_hours() -> bool:
    now = datetime.now()
    if now.weekday() >= 5:
        return False
    open_mins = MARKET_OPEN_HOUR * 60 + MARKET_OPEN_MIN
    close_mins = MARKET_CLOSE_HOUR * 60 + MARKET_CLOSE_MIN
    current_mins = now.hour * 60 + now.minute
    return open_mins <= current_mins <= close_mins


def pipeline(force: bool = False, extract: bool = False) -> bool:
    start_dt = datetime.now(timezone.utc)
    status_repo = PipelineStatusRepository(
        PIPELINE_STATUS_JSON,
        stale_after_seconds=(
            PIPELINE_RUN_TIMEOUT_SECONDS + PIPELINE_STALE_GRACE_SECONDS
        ),
    )
    requested_run_id = os.getenv("PIPELINE_RUN_ID")
    run_id = requested_run_id or str(uuid4())
    try:
        status_repo.start(
            run_id=run_id,
            expected_run_id=requested_run_id,
        )
    except (PipelineRunActiveError, PipelineRunInactiveError) as error:
        print(f"[{start_dt.strftime('%H:%M')}] {error} Skipping this invocation.")
        return False

    if not force and not is_market_hours():
        print(f"[{start_dt.strftime('%H:%M')}] Outside market hours — skipping")
        _append_run_log(
            {
                "run_id": run_id,
                "start": start_dt.isoformat(timespec="seconds"),
                "end": start_dt.isoformat(timespec="seconds"),
                "duration_seconds": 0,
                "scanned": 0,
                "analysed": 0,
                "buy_alerts": 0,
                "sell_alerts": 0,
                "actionable": 0,
                "sources": "",
                "source_health_json": "{}",
                "errors": "",
                "status": "skipped",
            }
        )
        status_repo.finish(PipelineState.SKIPPED, expected_run_id=run_id)
        return False

    print(f"\n{'=' * 50}")
    print(f"Pipeline run: {start_dt.strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * 50)

    errors: list[str] = []
    scanned = 0
    analysed = 0
    buy_alerts = 0
    sell_alerts = 0
    src_summary = ""
    source_health: dict[SourceName, SourceHealth] = {}
    terminal_state = PipelineState.FAILED
    temporary_artifacts: list[Path] = []
    last_analysis_update = 0.0

    def report_scanner(stage_name: str, state_name: str, count: int | None) -> None:
        stage = PipelineStage(stage_name)
        state = StageState(state_name)
        status_repo.transition(stage, state, expected_run_id=run_id)
        if stage is PipelineStage.ENRICHMENT and state is StageState.COMPLETE:
            status_repo.update_counts(scanned=count or 0, expected_run_id=run_id)

    def report_analysis(current: int, total: int) -> None:
        nonlocal last_analysis_update
        now = time.monotonic()
        if current != total and now - last_analysis_update < 1:
            return
        last_analysis_update = now
        status_repo.transition(
            PipelineStage.ANALYSIS,
            StageState.RUNNING,
            expected_run_id=run_id,
            current=current,
            total=total,
        )

    def report_step(event: PipelineStepEvent) -> None:
        mapping = {
            "scan": PipelineStage.SOURCES,
            "analyse": PipelineStage.ANALYSIS,
            "alert": PipelineStage.ALERTS,
        }
        stage = mapping[event.step_name]
        if event.phase == "started":
            status_repo.transition(stage, StageState.RUNNING, expected_run_id=run_id)
            return
        if event.step_name == "scan":
            # Empty scans do not enter ScannerAgent.scan_watchlist; make their
            # downstream scan substages explicit rather than leaving them pending.
            current = status_repo.load()
            for scan_stage in (
                PipelineStage.SOURCES,
                PipelineStage.MARKET_DATA,
                PipelineStage.ENRICHMENT,
            ):
                if current.stage(scan_stage).state in {
                    StageState.PENDING,
                    StageState.RUNNING,
                }:
                    status_repo.transition(
                        scan_stage,
                        StageState.COMPLETE,
                        expected_run_id=run_id,
                    )
            status_repo.update_counts(
                scanned=event.output_count or 0,
                expected_run_id=run_id,
            )
        elif event.step_name != "alert":
            status_repo.transition(stage, StageState.COMPLETE, expected_run_id=run_id)
            if event.step_name == "analyse":
                status_repo.update_counts(
                    analysed=event.output_count or 0,
                    expected_run_id=run_id,
                )

    try:
        status_repo.transition(
            PipelineStage.SOURCES,
            StageState.RUNNING,
            expected_run_id=run_id,
        )
        if extract:
            status_repo.transition(
                PipelineStage.SOURCES,
                StageState.RUNNING,
                expected_run_id=run_id,
                substage="extraction",
            )
            extractor = ExtractionAgent(name="ExtractionAgent")
            extractor.run()
            source_health.update(extractor.source_health)
        watchlist = load_watchlist()
        source_map = load_source_map()
        if not extract:
            source_health.update(_cached_extraction_health(source_map))

        scanner = ScannerAgent(name="ScannerAgent", progress_callback=report_scanner)
        analyst = AnalystAgent(progress_callback=report_analysis)
        alerter = AlertAgent(name="AlertAgent")
        # Inject the agents so the same alerter is reused below for portfolio
        # stop checks and the summary email.
        momentum = build_momentum_pipeline(scanner, analyst, alerter)

        alert_summary, trace = momentum.run_traced(watchlist, progress=report_step)
        scan_results = trace[0][1]
        analysis_results = trace[1][1]
        scanned = len(scan_results)
        analysed = len(analysis_results)
        source_health.update(scanner.source_health)
        for source in SourceName:
            source_health.setdefault(
                source,
                SourceHealth(
                    source=source,
                    state=SourceState.EMPTY,
                    count=0,
                    display_message="No records were available for this source.",
                ),
            )
        status_repo.update_source_health(source_health, expected_run_id=run_id)

        for r in analysis_results:
            st, ww = source_map.get(r.ticker, (False, False))
            r.in_stocktwits = st
            r.in_whale_wisdom = ww

        grouped: dict[str, list[dict]] = {k: [] for k in _SOURCE_COMMENTS}
        for item in scan_results:
            primary = item.source.split(",")[0] or "ww_extraction"
            grouped.setdefault(primary, []).append(item.model_dump())

        scan_payload: dict[str, object] = {
            "as_of": datetime.now().isoformat(timespec="seconds")
        }
        for src, items in grouped.items():
            scan_payload[src] = {
                "_comment": _SOURCE_COMMENTS.get(src, src),
                "results": items,
            }

        src_summary = ", ".join(
            f"{health.source.label}:{health.count} ({health.state.value})"
            for health in source_health.values()
        )
        print(f"      Scanned {scanned} tickers ({src_summary})")

        _trader = TraderAgent(name="TraderAgent")
        stock_map = {r.ticker: r for r in analysis_results}

        # Price every actual holding (not just this run's scanned watchlist)
        # cache-first, sharing the same price cache the web UI reads/writes,
        # so the email snapshot and the /portfolio route always agree.
        portfolio_service = PortfolioService(TraderService(_trader))
        held_tickers = [p.ticker for p in _trader.get_portfolio()]
        prices, display_info, gbpusd = portfolio_service.get_prices_for_holdings(
            held_tickers
        )
        positions = _trader.get_portfolio(prices, display_info)
        held = {p.ticker for p in positions}
        gbp_totals = PortfolioService.gbp_totals(positions, gbpusd)
        pf_value, pf_cost, _pf_pnl = gbp_totals
        if pf_value > 0:
            _append_portfolio_snapshot(pf_value, pf_cost)

        sell_alerts = alerter.check_portfolio_stops(positions, stock_map)
        # Trailing-stop detection runs after the hard-stop check so a ticker
        # already alerting this run is skipped (no duplicate email/digest), and
        # updates the persistent high-water-mark per held position (#82).
        sell_alerts += alerter.check_trailing_stops(positions)
        buy_alerts = alert_summary.buy_count
        status_repo.update_counts(
            actionable=buy_alerts + sell_alerts,
            expected_run_id=run_id,
        )

        # Deterministic sector-allocation + market-cycle narrative (#109,
        # Phase 1 — no API keys/network/LLM). Persisted below alongside
        # save_history so the next run's delta has a baseline, and cached to
        # disk so the web banner and this digest render the same snapshot
        # without recomputing.
        sector_history = load_sector_history()
        sector_snapshot = with_deltas(
            compute_sector_prevalence(analysis_results),
            latest_snapshot(sector_history),
        )
        ticker_sector = {r.ticker: r.sector or "Unknown" for r in analysis_results}
        portfolio_sector_weights = compute_portfolio_sector_weights(
            positions, ticker_sector, gbpusd
        )
        market_cycle = get_market_cycle_context()
        market_narrative = build_deterministic_narrative(
            sector_snapshot, portfolio_sector_weights, market_cycle
        )
        save_market_narrative(market_narrative)

        # TODO: Phase 3 consumes NewsContext here. Call
        # news_context.gather_news_context(top_sectors, av_client) with
        # top_sectors = [s.sector for s in sector_snapshot.shares[:3]] to
        # fetch this run's news/macro context, then use it (together with
        # sector_snapshot / portfolio_sector_weights / market_cycle) to build
        # a Claude-generated MarketNarrative in place of
        # build_deterministic_narrative above, populating `sources` from the
        # gathered NewsItems. Left un-wired in Phase 2 because GDELT can take
        # ~15-20s and Alpha Vantage's 25-calls/day quota is shared with
        # fundamentals lookups — both phases run only once Phase 3 actually
        # consumes the result.

        alerter.send_summary_email(
            positions, gbp_totals=gbp_totals, market_narrative=market_narrative
        )

        status_repo.transition(
            PipelineStage.ALERTS,
            StageState.COMPLETE,
            expected_run_id=run_id,
        )

        status_repo.transition(
            PipelineStage.EXPORT,
            StageState.RUNNING,
            expected_run_id=run_id,
        )

        if not analysis_results:
            errors.append("No usable analysis records were produced.")
            status_repo.transition(
                PipelineStage.EXPORT,
                StageState.SKIPPED,
                expected_run_id=run_id,
            )
            terminal_state = derive_pipeline_outcome(
                source_health,
                analysed=analysed,
                artifact_produced=False,
            )
            status_repo.finish(
                terminal_state,
                expected_run_id=run_id,
                scanned=scanned,
                analysed=analysed,
                actionable=buy_alerts + sell_alerts,
                error_summary="No usable analysis records were produced.",
                artifact_produced=False,
            )
            print("\nPipeline failed: no usable analysis records were produced.")
            return False

        def _temporary_output(path: str | Path) -> Path:
            """Return a run-scoped temp path next to *path*, tracked for cleanup."""
            target = Path(path)
            temporary = target.with_name(f".{target.stem}.{run_id}.tmp{target.suffix}")
            temporary.parent.mkdir(parents=True, exist_ok=True)
            temporary_artifacts.append(temporary)
            return temporary

        scan_temporary = _temporary_output(SCAN_OUTPUT)
        analysis_temporary = _temporary_output(ANALYSIS_OUTPUT)
        excel_temporary = _temporary_output(EXCEL_OUTPUT)

        with open(scan_temporary, "w", encoding="utf-8") as stream:
            json.dump(scan_payload, stream, indent=2)

        # Embed run ownership + generation time *inside* the artifact payload
        # so the single os.replace below atomically publishes both the data
        # and the fact that this run produced it. A separate follow-up write
        # to record ownership would leave a crash window where the file is
        # promoted but its owner is unrecorded (or vice versa) — see
        # app.schemas.analysis_artifact for the full rationale.
        analysis_payload = build_analysis_payload(
            [item.model_dump(mode="json") for item in analysis_results],
            run_id=run_id,
            generated_at=datetime.now(timezone.utc),
        )
        with open(analysis_temporary, "w", encoding="utf-8") as stream:
            json.dump(analysis_payload, stream, indent=2)

        current_tickers = [r.ticker for r in analysis_results]
        history = load_history()
        new = get_new_tickers(current_tickers, history)
        breakouts = get_fresh_breakouts(analysis_results, history)
        save_history(analysis_results, history)
        save_sector_snapshot(sector_snapshot, sector_history)

        for r in analysis_results:
            if r.analysis and r.ticker in breakouts:
                r.analysis.fresh_breakout = True

        if new:
            print(f"      New tickers this run:     {len(new)}")
        if breakouts:
            print(
                f"      Fresh breakouts this run: {len(breakouts)}"
                f"  ({', '.join(sorted(breakouts)[:8])}"
                f"{'...' if len(breakouts) > 8 else ''})"
            )

        write_excel(analysis_results, excel_temporary, held, new_tickers=new)
        # Publish scan/excel first; the analysis artifact is promoted last so
        # a failure partway through export can never leave a *new* analysis
        # file paired with stale scan/excel siblings.
        os.replace(scan_temporary, SCAN_OUTPUT)
        os.replace(excel_temporary, EXCEL_OUTPUT)
        os.replace(analysis_temporary, ANALYSIS_OUTPUT)
        status_repo.transition(
            PipelineStage.EXPORT,
            StageState.COMPLETE,
            expected_run_id=run_id,
        )
        terminal_state = derive_pipeline_outcome(
            source_health,
            analysed=analysed,
            artifact_produced=True,
        )
        status_repo.finish(
            terminal_state,
            expected_run_id=run_id,
            scanned=scanned,
            analysed=analysed,
            actionable=buy_alerts + sell_alerts,
            artifact_produced=True,
        )
        print(f"\nPipeline {terminal_state.value}.")

    except Exception as error:
        errors.append(traceback.format_exc())
        terminal_state = PipelineState.FAILED
        status_repo.finish(
            PipelineState.FAILED,
            expected_run_id=run_id,
            scanned=scanned,
            analysed=analysed,
            actionable=buy_alerts + sell_alerts,
            error_summary=str(error),
        )
        print(f"\n[ERROR] Pipeline failed:\n{errors[-1]}")

    finally:
        for temporary in temporary_artifacts:
            temporary.unlink(missing_ok=True)
        end_dt = datetime.now(timezone.utc)
        duration = round((end_dt - start_dt).total_seconds(), 1)
        log_entry = {
            "run_id": run_id,
            "start": start_dt.isoformat(timespec="seconds"),
            "end": end_dt.isoformat(timespec="seconds"),
            "duration_seconds": duration,
            "scanned": scanned,
            "analysed": analysed,
            "buy_alerts": buy_alerts,
            "sell_alerts": sell_alerts,
            "actionable": buy_alerts + sell_alerts,
            "sources": src_summary,
            "source_health_json": json.dumps(
                {
                    source.value: health.model_dump(mode="json")
                    for source, health in source_health.items()
                }
            ),
            "errors": " | ".join(errors) if errors else "",
            "status": (
                "error"
                if terminal_state is PipelineState.FAILED
                else terminal_state.value
            ),
        }
        _append_run_log(log_entry)
        _emit_run_notifications(
            terminal_state,
            source_health,
            run_id=run_id,
            scanned=scanned,
            analysed=analysed,
            buy_alerts=buy_alerts,
            sell_alerts=sell_alerts,
            duration=duration,
        )
        print(
            f"[log] run {log_entry['status'].upper()} | "
            f"{duration}s | scanned={scanned} analysed={analysed} "
            f"actionable={log_entry['actionable']} (buy={buy_alerts} sell={sell_alerts})"
        )
    return terminal_state in {PipelineState.COMPLETE, PipelineState.PARTIAL}


def main() -> None:
    parser = argparse.ArgumentParser(description="Momentum stock scanner")
    parser.add_argument(
        "--once", action="store_true", help="Run pipeline once and exit"
    )
    parser.add_argument(
        "--extract",
        action="store_true",
        help="Pull watchlist from WisdomWise instead of default",
    )
    parser.add_argument(
        "--interval",
        type=int,
        default=15,
        help="Polling interval in minutes (default: 15)",
    )
    args = parser.parse_args()

    if args.once:
        if not pipeline(force=True, extract=args.extract):
            raise SystemExit(1)
        return

    print(
        f"Scheduler started — running every {args.interval} minutes during market hours"
    )
    print("Press Ctrl+C to stop\n")

    scheduler = BlockingScheduler()
    scheduler.add_job(
        pipeline,
        CronTrigger(minute=f"*/{args.interval}"),
        id="momentum_scan",
        max_instances=1,
        coalesce=True,
    )

    pipeline(force=False)

    try:
        scheduler.start()
    except KeyboardInterrupt:
        print("\nScheduler stopped.")


if __name__ == "__main__":
    main()

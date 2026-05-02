"""
Orchestrator — wires the three agents together and schedules them.
Runs the MS Agent framework pipeline on market hours.
"""

import argparse
import csv
import json
import traceback
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger

from agents.analyst.analyst_agent import AnalystAgent, recommendation
from agents.alert.alert_agent import AlertAgent
from agents.extraction.extraction_agent import ExtractionAgent
from agents.trader.trader_agent import TraderAgent
from ms_agent_framework import AgentApp
from agents.scanner.scanner_agent import ScannerAgent, load_watchlist, load_source_map
from agents.scanner.scan_history import (
    get_fresh_breakouts,
    get_new_tickers,
    load_history,
    save_history,
)
from models import StockRecord


SCAN_OUTPUT = "agents/scanner/scan_results.json"
RUN_LOG = "pipeline_runs.csv"
PORTFOLIO_VALUE_LOG = "portfolio_value.csv"
_RUN_LOG_FIELDS = [
    "start", "end", "duration_seconds",
    "scanned", "analysed", "buy_alerts", "sell_alerts", "actionable",
    "sources", "status", "errors",
]
_SOURCE_COMMENTS: dict[str, str] = {
    "ww_extraction":   "WhaleWisdom heat map – institutional filer top holdings",
    "vcp_screener":    "Minervini pure VCP setup – S&P 500 screened via FMP API",
    "tv_screener":     "TradingView screener – Stage 2 pre-filter (price>SMA200, SMA50>SMA150, within 35% of 52w high)",
    "tv_screener_uk":  "TradingView screener – LSE Stage 2 pre-filter (price>SMA200, SMA50>SMA150, within 35% of 52w high)",
}
ANALYSIS_OUTPUT = "agents/analyst/analysis_results.json"
EXCEL_OUTPUT = "agents/analyst/analysis_results.xlsx"
MARKET_OPEN_HOUR = 9
MARKET_OPEN_MIN = 30
MARKET_CLOSE_HOUR = 16
MARKET_CLOSE_MIN = 0


def _append_portfolio_snapshot(total_value: float, total_cost: float) -> None:
    """Append one portfolio value snapshot (timestamped) to PORTFOLIO_VALUE_LOG."""
    log_path = Path(PORTFOLIO_VALUE_LOG)
    write_header = not log_path.exists() or log_path.stat().st_size == 0
    with open(log_path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["timestamp", "total_value", "total_cost"])
        if write_header:
            writer.writeheader()
        writer.writerow({
            "timestamp": datetime.now(timezone.utc).isoformat(timespec="minutes"),
            "total_value": round(total_value, 2),
            "total_cost": round(total_cost, 2),
        })


def _append_run_log(entry: dict) -> None:
    """Append one run record to the CSV log, writing the header if the file is new."""
    log_path = Path(RUN_LOG)
    write_header = not log_path.exists() or log_path.stat().st_size == 0
    with open(log_path, "a", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_RUN_LOG_FIELDS)
        if write_header:
            writer.writeheader()
        writer.writerow({k: entry.get(k, "") for k in _RUN_LOG_FIELDS})

_SEPA_LABELS = {
    "above_150_200":        "P>SMA150&200",
    "sma150_above_200":     "SMA150>200",
    "sma200_rising":        "SMA200 Rising",
    "sma50_above_150_200":  "SMA50>150&200",
    "above_25pct_of_low":   ">=25% abv Low",
    "within_25pct_of_high": "<=25% frm High",
    "rs_leader":            "RS Leader",
    "above_sma50":          "P>SMA50",
}
_SEPA_KEYS = list(_SEPA_LABELS.keys())

_HEADERS = [
    "Ticker", "Market", "Source", "StockTwits", "Whale Wisdom", "Score", "CANSLIM", "Momentum", "Stage", "Zone",
    "Prev Entry", "Next Entry", "Stop", "Risk %", "1R Target", "2R Target", "3R Target", "R:R", "Price", "P/E", "RSI", "Rel Vol",
    "% 52w High", "% Chg Week", "Rel Str vs SPY",
    "EPS Growth", "ROE", "Inst Ownership %", "Inst Count", "WW Buyers", "WW Sellers", "WW Net",
    "Congress Buys", "Congress Sells", "Congress Net",
    "Senate Buys", "Senate Sells", "Senate Net",
    "SPY Uptrend",
    "SMA Stack", "SMA50", "Near High", "RSI Zone", "Vol Zone",
    "SEPA", *[f"SEPA: {v}" for v in _SEPA_LABELS.values()],
    "Sector",
    "VCP Breakout", "MYB Breakout", "MYB Pivot",
    "Recommendation", "As Of", "Summary",
]

# Derive column positions dynamically so they stay correct as columns are added.
_SIGNAL_COL_START = _HEADERS.index("SMA Stack") + 1   # 1-indexed
_SIGNAL_COL_END   = _HEADERS.index("Vol Zone") + 1
_SEPA_COL_START   = _HEADERS.index("SEPA") + 1
_SEPA_COL_END     = _SEPA_COL_START + len(_SEPA_KEYS)  # summary + 8 conditions (exclusive end = summary+8)

_SCORE_FILLS = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),  # green  ≥8
    "mid":    PatternFill("solid", fgColor="FFEB9C"),  # yellow 6–7
    "low":    PatternFill("solid", fgColor="FFC7CE"),  # red    ≤5
}
_SIGNAL_FILLS = {
    "+": PatternFill("solid", fgColor="C6EFCE"),  # green  — positive for buying
    "-": PatternFill("solid", fgColor="FFC7CE"),  # red    — negative / caution
    "~": PatternFill("solid", fgColor="FFEB9C"),  # yellow — neutral
}
_HEADER_FILL = PatternFill("solid", fgColor="2F5496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_NEAR_ENTRY_FILL = PatternFill("solid", fgColor="E2EFDA")


def _yahoo_url(ticker: str) -> str:
    return f"https://finance.yahoo.com/quote/{ticker.replace('.', '-')}"


def _score_fill(score: int) -> PatternFill:
    if score >= 8:
        return _SCORE_FILLS["high"]
    if score >= 6:
        return _SCORE_FILLS["mid"]
    return _SCORE_FILLS["low"]


def _pct(value: float | None) -> str:
    return f"{value * 100:.1f}%" if value is not None else ""


def _signals(r: StockRecord) -> list[str]:
    """Return 5 buy-signal indicators (+/-/~) from scan data.

    These correspond to the conditions previously used in the heuristic score
    and are shown as coloured columns in the Excel output.
    """
    sma150 = r.sma150
    sma200 = r.sma200
    sma50 = r.sma50
    rsi = r.rsi14
    pct = r.pct_from_52w_high
    vol = r.rel_volume

    # SMA Stack: price > SMA150 > SMA200 → bullish macro structure
    if sma150 and sma200 and r.price > sma150 > sma200:
        sma_stack = "+"
    elif sma150 and sma200 and r.price < sma150 and r.price < sma200:
        sma_stack = "-"
    else:
        sma_stack = "~"

    # SMA50: short-term trend
    sma50_sig = "+" if sma50 and r.price > sma50 else "-"

    # Near High: -15% to -1% is the ideal consolidation zone; < -30% is too extended
    if pct is not None and -15 <= pct <= -1:
        near_high = "+"
    elif pct is not None and pct < -30:
        near_high = "-"
    else:
        near_high = "~"

    # RSI Zone: 50-80 healthy, >80 overbought, <40 weak, 40-50 neutral
    if rsi is not None and 50 <= rsi <= 80:
        rsi_zone = "+"
    elif rsi is not None and (rsi > 80 or rsi < 40):
        rsi_zone = "-"
    else:
        rsi_zone = "~"

    # Vol Zone: elevated volume signals institutional participation
    if vol >= 1.5:
        vol_zone = "+"
    elif vol < 0.7:
        vol_zone = "-"
    else:
        vol_zone = "~"

    return [sma_stack, sma50_sig, near_high, rsi_zone, vol_zone]


def _record_to_row(
    r: StockRecord, portfolio_tickers: set[str] | None = None
) -> list[object]:
    a = r.analysis
    if not a:
        return [r.ticker] + [""] * (len(_HEADERS) - 1)
    risk = (
        f"{(a.entry_price - a.stop_loss) / a.entry_price * 100:.1f}%"
        if a.entry_price and a.stop_loss
        else ""
    )
    tmpl = a.sepa_template or {}
    sepa_vals = [tmpl.get(k, False) for k in _SEPA_KEYS]
    sepa_summary = f"{sum(sepa_vals)}/8"

    rec = recommendation(a)
    if rec == "Avoid" and portfolio_tickers and r.ticker in portfolio_tickers:
        rec = "SELL"

    return [
        r.ticker,
        "UK" if r.ticker.endswith(".L") else "US",
        r.source,
        "Y" if r.in_stocktwits else "",
        "Y" if r.in_whale_wisdom else "",
        f"{a.score}/10",
        f"{a.canslim.total}/14" if a.canslim else "",
        f"{a.momentum.total}/14" if a.momentum else "",
        a.stage,
        a.entry_zone,
        a.prev_entry_price,
        a.entry_price,
        a.stop_loss,
        risk,
        a.r_multiples.get("1.0R") if a.r_multiples else "",
        a.r_multiples.get("2.0R") if a.r_multiples else "",
        a.r_multiples.get("3.0R") if a.r_multiples else "",
        f"{a.reward_risk_ratio:.1f}x" if a.reward_risk_ratio else "",
        r.price,
        round(r.pe_ratio, 1) if r.pe_ratio else "",
        r.rsi14,
        r.rel_volume,
        r.pct_from_52w_high,
        r.pct_change_week,
        r.rel_strength_vs_spy,
        _pct(r.eps_growth),
        _pct(r.roe),
        _pct(r.inst_ownership_pct),
        r.inst_count if r.inst_count is not None else "",
        r.funds_buying if r.funds_buying is not None else "",
        r.funds_selling if r.funds_selling is not None else "",
        r.funds_net if r.funds_net is not None else "",
        r.congress_buys if r.congress_buys is not None else "",
        r.congress_sells if r.congress_sells is not None else "",
        (r.congress_buys or 0) - (r.congress_sells or 0)
        if r.congress_buys is not None or r.congress_sells is not None else "",
        r.senate_buys if r.senate_buys is not None else "",
        r.senate_sells if r.senate_sells is not None else "",
        (r.senate_buys or 0) - (r.senate_sells or 0)
        if r.senate_buys is not None or r.senate_sells is not None else "",
        "Y" if r.spy_uptrend else "N",
        *_signals(r),
        sepa_summary,
        *["Y" if v else "N" for v in sepa_vals],
        r.sector or "",
        "Y" if a.fresh_breakout else "",
        "Y" if a.multiyear_breakout else "",
        a.multiyear_pivot if a.multiyear_pivot else "",
        rec,
        r.as_of,
        a.summary,
    ]


def _apply_header(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    ws.append(_HEADERS)
    for col, _ in enumerate(_HEADERS, 1):
        cell = ws.cell(1, col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _set_col_widths(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    widths = {
        "A": 8,   # Ticker
        "B": 18,  # Source
        "C": 11,  # StockTwits
        "D": 12,  # Whale Wisdom
        "E": 8,   # Score
        "F": 10,  # CANSLIM
        "G": 10,  # Momentum
        "H": 10,  # Stage
        "I": 8,   # Zone
        "J": 11,  # Prev Entry
        "K": 11,  # Next Entry
        "L": 10,  # Stop
        "M": 8,   # Risk %
        "N": 10,  # 1R Target
        "O": 10,  # 2R Target
        "P": 10,  # 3R Target
        "Q": 7,   # R:R
        "R": 10,  # Price
        "S": 7,   # P/E
        "T": 7,   # RSI
        "U": 8,   # Rel Vol
        "V": 11,  # % 52w High
        "V": 11,  # % Chg Week
        "W": 14,  # Rel Str vs SPY
        "X": 11,  # EPS Growth
        "Y": 8,   # ROE
        "Z": 15,  # Inst Ownership
        "AA": 10, # Inst Count
        "AB": 9,  # WW Buyers
        "AC": 9,  # WW Sellers
        "AD": 8,  # WW Net
        "AE": 12, # Congress Buys
        "AF": 12, # Congress Sells
        "AG": 10, # Congress Net
        "AH": 10, # Senate Buys
        "AI": 10, # Senate Sells
        "AJ": 10, # Senate Net
        "AK": 10, # SPY Uptrend
        "AL": 9,  # SMA Stack
        "AM": 7,  # SMA50
        "AN": 9,  # Near High
        "AO": 8,  # RSI Zone
        "AP": 8,  # Vol Zone
        "AQ": 7,  # SEPA summary
        "AR": 13, "AS": 13, "AT": 13, "AU": 13,  # SEPA conditions 1-4
        "AV": 13, "AW": 13, "AX": 13, "AY": 13,  # SEPA conditions 5-8
        "AZ": 12, # Sector
        "BA": 12, # Recommendation
        "BB": 12, # As Of
        "BC": 50, # Summary
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _format_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    data_row: int,
    r: StockRecord,
) -> None:
    """Apply fills, hyperlink, and alignment for a single data row."""
    score = r.analysis.score if r.analysis else 0
    fill = _score_fill(score)
    link_cell = ws.cell(data_row, 1)
    assert isinstance(link_cell, Cell)
    link_cell.hyperlink = Hyperlink(ref="", target=_yahoo_url(r.ticker))
    link_cell.font = Font(color="0563C1", underline="single")
    link_cell.fill = fill
    ws.cell(data_row, 4).fill = fill
    if r.analysis and r.analysis.entry_zone in ("broken_out", "approaching"):
        ws.cell(data_row, 8).fill = _NEAR_ENTRY_FILL

    for col in range(_SIGNAL_COL_START, _SIGNAL_COL_END + 1):
        cell = ws.cell(data_row, col)
        if cell.value in _SIGNAL_FILLS:
            cell.fill = _SIGNAL_FILLS[cell.value]
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(_SEPA_COL_START + 1, _SEPA_COL_END + 1):
        cell = ws.cell(data_row, col)
        if cell.value == "Y":
            cell.fill = _SIGNAL_FILLS["+"]
        elif cell.value == "N":
            cell.fill = _SIGNAL_FILLS["-"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sepa_cell = ws.cell(data_row, _SEPA_COL_START)
    try:
        count = int(str(sepa_cell.value).split("/")[0])
    except (ValueError, AttributeError):
        count = 0
    if count == 8:
        sepa_fill = _SIGNAL_FILLS["+"]
    elif count >= 6:
        sepa_fill = _SIGNAL_FILLS["~"]
    else:
        sepa_fill = _SIGNAL_FILLS["-"]
    sepa_cell.fill = sepa_fill
    sepa_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_name in ("VCP Breakout", "MYB Breakout"):
        col = _HEADERS.index(col_name) + 1
        cell = ws.cell(data_row, col)
        if cell.value == "Y":
            cell.fill = _SIGNAL_FILLS["+"]
        cell.alignment = Alignment(horizontal="center", vertical="center")

    market_col = _HEADERS.index("Market") + 1
    market_cell = ws.cell(data_row, market_col)
    if market_cell.value == "UK":
        market_cell.fill = PatternFill("solid", fgColor="DBEAFE")  # pale blue
    market_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, len(_HEADERS) + 1):
        ws.cell(data_row, col).alignment = Alignment(vertical="center")


_NEW_TICKER_FILL = PatternFill("solid", fgColor="FFF2CC")  # pale yellow — new tickers
_BREAKOUT_FILL   = PatternFill("solid", fgColor="FFE0B2")  # pale orange — fresh breakouts


def _write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    records: list[StockRecord],
    held: set[str],
    ranked: bool = False,
) -> None:
    """Populate a worksheet with header + formatted data rows."""
    _apply_header(ws)
    _set_col_widths(ws)
    for rank, r in enumerate(records, 1):
        row = _record_to_row(r, held)
        if ranked:
            row[0] = f"{rank}. {r.ticker}"
        ws.append(row)
        _format_row(ws, ws.max_row, r)


def write_excel(
    records: list[StockRecord],
    path: str,
    portfolio_tickers: set[str] | None = None,
    new_tickers: set[str] | None = None,
) -> None:
    """Write analysis results to a formatted Excel workbook."""
    held = portfolio_tickers or set()
    new = new_tickers or set()
    wb = openpyxl.Workbook()

    # --- Sheet 1: All results ---
    ws_all = wb.active
    ws_all.title = "All Results"
    _write_sheet(ws_all, records, held)

    # Highlight new tickers with a pale-yellow row tint on Sheet 1
    for row_idx in range(2, ws_all.max_row + 1):
        ticker_cell = ws_all.cell(row_idx, 1)
        raw = str(ticker_cell.value or "").lstrip("0123456789. ")
        if raw in new:
            for col in range(1, len(_HEADERS) + 1):
                cell = ws_all.cell(row_idx, col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _NEW_TICKER_FILL

    # --- Sheet 2: New Tickers ---
    new_records = [r for r in records if r.ticker in new]
    if new_records:
        ws_new = wb.create_sheet("New Tickers")
        _write_sheet(ws_new, new_records, held)
        # All rows on this sheet get the yellow tint
        for row_idx in range(2, ws_new.max_row + 1):
            for col in range(1, len(_HEADERS) + 1):
                cell = ws_new.cell(row_idx, col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _NEW_TICKER_FILL

    # --- Sheet 3: Fresh Breakouts ---
    breakout_records = [r for r in records if r.analysis and r.analysis.fresh_breakout]
    if breakout_records:
        ws_bo = wb.create_sheet("Breakouts")
        _write_sheet(ws_bo, breakout_records, held)
        for row_idx in range(2, ws_bo.max_row + 1):
            for col in range(1, len(_HEADERS) + 1):
                cell = ws_bo.cell(row_idx, col)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = _BREAKOUT_FILL

    # --- Sheet 4: Top 20 ---
    top20 = [r for r in records if r.analysis][:20]
    ws_top = wb.create_sheet("Top 20")
    _write_sheet(ws_top, top20, held, ranked=True)

    wb.save(path)
    parts = [f"{len(records)} records"]
    if new_records:
        parts.append(f"{len(new_records)} new")
    if breakout_records:
        parts.append(f"{len(breakout_records)} breakouts")
    print(f"      Excel saved -> {path}  ({', '.join(parts)})")


def is_market_hours() -> bool:
    now = datetime.now()
    if now.weekday() >= 5:
        return False
    open_mins = MARKET_OPEN_HOUR * 60 + MARKET_OPEN_MIN
    close_mins = MARKET_CLOSE_HOUR * 60 + MARKET_CLOSE_MIN
    current_mins = now.hour * 60 + now.minute
    return open_mins <= current_mins <= close_mins


def pipeline(force: bool = False, extract: bool = False) -> None:
    if not force and not is_market_hours():
        print(f"[{datetime.now().strftime('%H:%M')}] Outside market hours — skipping")
        return

    start_dt = datetime.now(timezone.utc)
    print(f"\n{'='*50}")
    print(f"Pipeline run: {start_dt.strftime('%Y-%m-%d %H:%M UTC')}")
    print("="*50)

    errors: list[str] = []
    scanned = 0
    analysed = 0
    buy_alerts = 0
    sell_alerts = 0
    src_summary = ""

    try:
        if extract:
            ExtractionAgent(name="ExtractionAgent").run()
        watchlist = load_watchlist()
        source_map = load_source_map()

        scanner = ScannerAgent(name="ScannerAgent")
        analyst = AnalystAgent()
        alerter = AlertAgent(name="AlertAgent")
        app = AgentApp(name="MomentumStockAgent")
        app.add_agent(scanner)
        app.add_agent(analyst)
        app.add_agent(alerter)

        _, intermediates = app.execute_with_intermediates(watchlist)
        scan_results, analysis_results = intermediates
        scanned = len(scan_results)
        analysed = len(analysis_results)

        for r in analysis_results:
            st, ww = source_map.get(r.ticker, (False, False))
            r.in_stocktwits = st
            r.in_whale_wisdom = ww

        grouped: dict[str, list[dict]] = {k: [] for k in _SOURCE_COMMENTS}
        for item in scan_results:
            primary = item.source.split(",")[0] or "ww_extraction"
            grouped.setdefault(primary, []).append(item.model_dump())

        scan_payload: dict[str, object] = {"as_of": datetime.now().isoformat(timespec="seconds")}
        for src, items in grouped.items():
            scan_payload[src] = {"_comment": _SOURCE_COMMENTS.get(src, src), "results": items}

        with open(SCAN_OUTPUT, "w", encoding="utf-8") as stream:
            json.dump(scan_payload, stream, indent=2)

        src_summary = ", ".join(
            f"{k}:{len(v['results'])}"  # type: ignore[index]
            for k, v in scan_payload.items()
            if isinstance(v, dict)
        )
        print(f"      Scanned {scanned} tickers ({src_summary})")

        with open(ANALYSIS_OUTPUT, "w", encoding="utf-8") as stream:
            json.dump([item.model_dump() for item in analysis_results], stream, indent=2)

        _trader = TraderAgent(name="TraderAgent")
        stock_map = {r.ticker: r for r in analysis_results}
        prices = {r.ticker: r.price for r in analysis_results}
        positions = _trader.get_portfolio(prices)
        held = {p.ticker for p in positions}
        pf_value = sum(p.current_value for p in positions if p.current_value)
        pf_cost = sum(p.total_cost for p in positions)
        if pf_value > 0:
            _append_portfolio_snapshot(pf_value, pf_cost)

        sell_alerts = alerter.check_portfolio_stops(positions, stock_map)
        buy_alerts = len(alerter._buy_alerts)
        alerter.send_summary_email(positions)

        current_tickers = [r.ticker for r in analysis_results]
        history = load_history()
        new = get_new_tickers(current_tickers, history)
        breakouts = get_fresh_breakouts(analysis_results, history)
        save_history(analysis_results, history)

        for r in analysis_results:
            if r.analysis and r.ticker in breakouts:
                r.analysis.fresh_breakout = True

        if new:
            print(f"      New tickers this run:     {len(new)}")
        if breakouts:
            print(f"      Fresh breakouts this run: {len(breakouts)}"
                  f"  ({', '.join(sorted(breakouts)[:8])}"
                  f"{'...' if len(breakouts) > 8 else ''})")

        write_excel(analysis_results, EXCEL_OUTPUT, held, new_tickers=new)
        print("\nPipeline complete.")

    except Exception:
        errors.append(traceback.format_exc())
        print(f"\n[ERROR] Pipeline failed:\n{errors[-1]}")

    finally:
        end_dt = datetime.now(timezone.utc)
        duration = round((end_dt - start_dt).total_seconds(), 1)
        log_entry = {
            "start": start_dt.isoformat(timespec="seconds"),
            "end": end_dt.isoformat(timespec="seconds"),
            "duration_seconds": duration,
            "scanned": scanned,
            "analysed": analysed,
            "buy_alerts": buy_alerts,
            "sell_alerts": sell_alerts,
            "actionable": buy_alerts + sell_alerts,
            "sources": src_summary,
            "errors": " | ".join(errors) if errors else "",
            "status": "error" if errors else "ok",
        }
        _append_run_log(log_entry)
        print(
            f"[log] run {log_entry['status'].upper()} | "
            f"{duration}s | scanned={scanned} analysed={analysed} "
            f"actionable={log_entry['actionable']} (buy={buy_alerts} sell={sell_alerts})"
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Momentum stock scanner")
    parser.add_argument("--once", action="store_true", help="Run pipeline once and exit")
    parser.add_argument("--extract", action="store_true", help="Pull watchlist from WisdomWise instead of default")
    parser.add_argument(
        "--interval",
        type=int,
        default=15,
        help="Polling interval in minutes (default: 15)",
    )
    args = parser.parse_args()

    if args.once:
        pipeline(force=True, extract=args.extract)
        return

    print(f"Scheduler started — running every {args.interval} minutes during market hours")
    print("Press Ctrl+C to stop\n")

    scheduler = BlockingScheduler()
    scheduler.add_job(
        pipeline,
        CronTrigger(minute=f"*/{args.interval}"),
        id="momentum_scan",
        max_instances=1,
        coalesce=True,
    )

    pipeline(force=False)

    try:
        scheduler.start()
    except KeyboardInterrupt:
        print("\nScheduler stopped.")


if __name__ == "__main__":
    main()
